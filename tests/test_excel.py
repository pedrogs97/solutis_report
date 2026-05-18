import openpyxl

from infrastructure.excel.supplier_evaluation import SupplierEvaluationExcelGenerator


def test_generate_excel():
    generator = SupplierEvaluationExcelGenerator()
    data = [
        {
            "trade_name": "Fornecedor Teste",
            "tax_id": "12.345.678/0001-90",
            "final_score": "95.50",
            "period": "1º Quadrimestre",
            "evaluation_year": 2024,
            "evaluator_name": "João Silva",
            "evaluation_date": "2024-05-16",
        }
    ]

    excel_bytes = generator.generate(data)
    assert excel_bytes is not None

    # Load workbook from bytes to verify
    wb = openpyxl.load_workbook(excel_bytes)
    ws = wb.active

    # Check headers
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Nome",
        "CNPJ",
        "Pontuação",
        "Período",
        "Ano",
        "Avaliador",
        "Data da Avaliação",
    ]

    # Check data row
    row_data = [cell.value for cell in ws[2]]
    assert row_data == [
        "Fornecedor Teste",
        "12.345.678/0001-90",
        "95.50",
        "1º Quadrimestre",
        2024,
        "João Silva",
        "2024-05-16",
    ]
